import argparse
import base64
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


COUNTRY_TOKENS = [
    "spain",
    "espana",
    "espana",
    "espa",
    "esp",
    "espagne",
    "sp",
    "switzerland",
    "swiss",
    "suiza",
    "portugal",
    "france",
    "italy",
    "italia",
    "germany",
    "deutschland",
    "argentina",
    "uruguay",
]

BASE_DIR = Path(__file__).resolve().parents[1]
LOGO_PATH = BASE_DIR / "assets" / "logo.png"


def normalize_name(value: str) -> str:
    if not value:
        return ""
    name = str(value).strip()
    for token in COUNTRY_TOKENS:
        escaped = re.escape(token)
        name = re.sub(rf"\b{escaped}\b", "", name, flags=re.IGNORECASE)
        name = re.sub(rf"{escaped}\s*$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"[\u00A9]", "", name)
    name = re.sub(r"[\W_]+", " ", name, flags=re.UNICODE)
    name = name.lower().strip()
    name = unicodedata.normalize("NFD", name)
    name = "".join(ch for ch in name if unicodedata.category(ch) != "Mn")
    name = re.sub(r"\s+", " ", name).strip()
    return name


def read_scores(ws):
    scores = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1] if len(row) > 1 else None
        golpes = row[2] if len(row) > 2 else None
        if not name or golpes is None:
            continue
        try:
            golpes_int = int(golpes)
        except (TypeError, ValueError):
            continue
        key = normalize_name(str(name))
        if key and key not in scores:
            scores[key] = (str(name).strip(), golpes_int)
    return scores


def read_teams(ws):
    teams = {}
    header = [cell.value for cell in ws[1]]
    for col_idx, team_name in enumerate(header, start=1):
        if not team_name:
            continue
        team = str(team_name).strip()
        players = []
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value:
                players.append(str(value).strip())
        if players:
            teams[team] = players
    return teams


def find_score(player, scores, score_keys):
    key = normalize_name(player)
    if not key:
        return None
    if key in scores:
        return scores[key]
    matches = [k for k in score_keys if k in key or key in k]
    if matches:
        matches.sort(key=len)
        return scores[matches[0]]
    return None


def compute_results(teams, scores):
    score_keys = list(scores.keys())
    results = []
    for team, players in teams.items():
        found = []
        for player in players:
            match = find_score(player, scores, score_keys)
            if match is not None:
                found.append((match[0], match[1]))
        found.sort(key=lambda item: item[1])
        top4 = found[:4]
        total = sum(score for _, score in top4)
        results.append(
            {
                "Equipo": team,
                "TotalGolpes": total,
                "JugadoresPuntuaron": ", ".join(name for name, _ in top4),
                "GolpesPuntuaron": ", ".join(str(score) for _, score in top4),
            }
        )
    results.sort(key=lambda item: item["TotalGolpes"])
    return results


def compute_team_points(results, stage_count):
    points_table = [
        100,
        96,
        92,
        88,
        84,
        80,
        77,
        74,
        71,
        69,
        67,
        65,
        63,
        61,
        59,
        57,
        55,
        53,
        51,
        49,
    ]
    rows = []
    for idx, item in enumerate(results):
        stage_points = [0] * stage_count
        if idx < len(points_table):
            stage_points[0] = points_table[idx]
        rows.append(
            {
                "Equipo": item["Equipo"],
                "Etapas": stage_points,
                "Total": sum(stage_points),
            }
        )
    return rows


def build_player_groups(teams, scores, stage_count):
    score_keys = list(scores.keys())
    groups = []
    for team, players in teams.items():
        found = []
        for player in players:
            match = find_score(player, scores, score_keys)
            if match is not None:
                found.append((player, match[0], match[1]))
        found.sort(key=lambda item: item[2])
        top4 = {normalize_name(item[0]) for item in found[:4]}

        rows = []
        for player in players:
            match = find_score(player, scores, score_keys)
            etapa_scores = [180] * stage_count
            etapa_scored = [False] * stage_count
            if match is not None:
                etapa_scores[0] = match[1]
                etapa_scored[0] = normalize_name(player) in top4
            rows.append(
                {
                    "Jugador": player,
                    "Etapas": etapa_scores,
                    "Scored": etapa_scored,
                }
            )
        groups.append({"Equipo": team, "Rows": rows})
    return groups

def build_logo_data_uri(path):
    if not path or not path.exists():
        return ""
    data = path.read_bytes()
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:image/png;base64,{encoded}"

def write_clasificacion_sheet(wb, results, sheet_name="Clasificacion equipos"):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    ws.append(["Equipo", "TotalGolpes", "JugadoresPuntuaron", "GolpesPuntuaron"])
    for item in results:
        ws.append(
            [
                item["Equipo"],
                item["TotalGolpes"],
                item["JugadoresPuntuaron"],
                item["GolpesPuntuaron"],
            ]
        )


def build_html(results, output_path, classification_rows, player_groups, stage_names, logo_path):
    cards = []
    for idx, item in enumerate(results, start=1):
        players = []
        if item["JugadoresPuntuaron"] and item["GolpesPuntuaron"]:
            names = [n.strip() for n in item["JugadoresPuntuaron"].split(",")]
            points = [p.strip() for p in item["GolpesPuntuaron"].split(",")]
            for i, name in enumerate(names):
                score = points[i] if i < len(points) else ""
                players.append(
                    f"<li><span class='player'>{name}</span><span class='score'>{score}</span></li>"
                )
        medal = ""
        if idx == 1:
            medal = "gold"
        elif idx == 2:
            medal = "silver"
        elif idx == 3:
            medal = "bronze"
        cards.append(
            f"""
      <article class=\"card {medal}\">
        <div class=\"rank\">#{idx}</div>
        <div class=\"team\">{item['Equipo']}</div>
        <div class=\"total\">{item['TotalGolpes']} golpes</div>
        <ul class=\"players\">{''.join(players)}</ul>
      </article>
    """
        )

    team_colors = [
        "#f8e1b8",
        "#d6ecf4",
        "#f4d6e0",
        "#e1f1d2",
        "#f1e1c4",
        "#dfe2f6",
        "#f6e0d2",
        "#d2f0e9",
    ]

    logo_data = build_logo_data_uri(logo_path)
    logo_html = (
        f"<img class='logo' src='{logo_data}' alt='Logo' />" if logo_data else ""
    )
    html = f"""<!doctype html>
<html lang=\"es\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>Clasificacion Equipos - Etapa 1 2026</title>
  <link rel=\"preconnect\" href=\"https://fonts.googleapis.com\">
  <link rel=\"preconnect\" href=\"https://fonts.gstatic.com\" crossorigin>
  <link href=\"https://fonts.googleapis.com/css2?family=Fraunces:wght@600;700&family=IBM+Plex+Sans:wght@400;500;600&display=swap\" rel=\"stylesheet\">
  <style>
    :root {{
      --bg1: #f7f1e3;
      --bg2: #e9f4f5;
      --ink: #1c2329;
      --muted: #5b6a75;
      --accent: #b23a48;
      --gold: #c8a23d;
      --silver: #8a97a6;
      --bronze: #b26a4c;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: 'IBM Plex Sans', sans-serif;
      color: var(--ink);
      background: radial-gradient(1200px 600px at 10% -10%, #fff 0%, transparent 70%),
                  linear-gradient(135deg, var(--bg1), var(--bg2));
    }}
        header {{
            padding: 32px 24px 8px;
            text-align: center;
        }}
        .headline {{
            font-family: 'Fraunces', serif;
            font-size: clamp(22px, 3vw, 30px);
            margin: 0 0 8px;
            text-transform: uppercase;
            letter-spacing: 2px;
            background: linear-gradient(90deg, #c60b1e 0%, #ffc400 50%, #c60b1e 100%);
            -webkit-background-clip: text;
            background-clip: text;
            color: transparent;
            display: inline-block;
            padding: 4px 12px;
            border-bottom: 3px solid #ffc400;
        }}
        .logo {{
            width: 120px;
            height: auto;
            display: block;
            margin: 0 auto 12px;
            filter: drop-shadow(0 6px 12px rgba(0,0,0,0.18));
        }}
    h1 {{
      font-family: 'Fraunces', serif;
      font-size: clamp(28px, 4vw, 44px);
      margin: 0 0 6px;
      letter-spacing: 0.5px;
    }}
    .subtitle {{
      color: var(--muted);
      font-size: 14px;
      text-transform: uppercase;
      letter-spacing: 2px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
      gap: 18px;
      padding: 24px;
      max-width: 1100px;
      margin: 0 auto 32px;
    }}
    .card {{
      background: #ffffffcc;
      border-radius: 18px;
      padding: 18px 20px 16px;
      box-shadow: 0 8px 24px rgba(0,0,0,0.08);
      border: 1px solid rgba(0,0,0,0.06);
      position: relative;
      overflow: hidden;
    }}
    .card::after {{
      content: "";
      position: absolute;
      inset: auto -20% -40% auto;
      width: 220px;
      height: 220px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(178,58,72,0.12), transparent 70%);
    }}
    .rank {{
      font-family: 'Fraunces', serif;
      font-size: 22px;
      color: var(--accent);
    }}
    .team {{
      font-weight: 600;
      font-size: 18px;
      margin-top: 6px;
    }}
    .total {{
      font-size: 15px;
      color: var(--muted);
      margin-top: 10px;
      text-align: center;
      font-weight: 600;
      letter-spacing: 0.6px;
      text-transform: uppercase;
      padding: 8px 10px;
      border-radius: 999px;
      background: rgba(28,35,41,0.06);
      box-shadow: inset 0 0 0 1px rgba(28,35,41,0.08);
    }}
    .players {{
      list-style: none;
      padding: 10px 0 0;
      margin: 0;
    }}
    .players li {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      padding: 6px 0;
      border-bottom: 1px dashed rgba(0,0,0,0.08);
      font-size: 14px;
    }}
    .players li:last-child {{ border-bottom: none; }}
    .score {{
      font-variant-numeric: tabular-nums;
      color: var(--accent);
      font-weight: 600;
    }}
    .gold {{
      border: 1px solid rgba(200,162,61,0.55);
      background: linear-gradient(135deg, rgba(200,162,61,0.35), rgba(255,255,255,0.75));
    }}
    .gold .rank {{ color: var(--gold); }}
    .gold .total {{
      background: linear-gradient(135deg, rgba(200,162,61,0.45), rgba(255,255,255,0.9));
      color: #6b4b00;
      box-shadow: 0 6px 14px rgba(200,162,61,0.35);
    }}
    .silver {{
      border: 1px solid rgba(138,151,166,0.55);
      background: linear-gradient(135deg, rgba(138,151,166,0.35), rgba(255,255,255,0.75));
    }}
    .silver .rank {{ color: var(--silver); }}
    .silver .total {{
      background: linear-gradient(135deg, rgba(138,151,166,0.45), rgba(255,255,255,0.9));
      color: #405060;
      box-shadow: 0 6px 14px rgba(138,151,166,0.35);
    }}
    .bronze {{
      border: 1px solid rgba(178,106,76,0.55);
      background: linear-gradient(135deg, rgba(178,106,76,0.35), rgba(255,255,255,0.75));
    }}
    .bronze .rank {{ color: var(--bronze); }}
    .bronze .total {{
      background: linear-gradient(135deg, rgba(178,106,76,0.45), rgba(255,255,255,0.9));
      color: #6a3924;
      box-shadow: 0 6px 14px rgba(178,106,76,0.35);
    }}
        .table-section {{
            max-width: 1100px;
            margin: 0 auto 32px;
            padding: 0 24px 32px;
        }}
        .table-card {{
            background: #ffffffd9;
            border-radius: 16px;
            padding: 16px;
            margin-top: 20px;
            box-shadow: 0 10px 26px rgba(0,0,0,0.08);
            border: 1px solid rgba(0,0,0,0.06);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        th, td {{
            padding: 8px 10px;
            border-bottom: 1px solid rgba(0,0,0,0.08);
            text-align: center;
        }}
        th {{
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: var(--muted);
            text-align: center;
        }}
        td.team-cell, th.team-cell {{
            text-align: left;
            font-weight: 600;
        }}
        .scored {{
            background: #c6efce;
        }}
        .team-block {{
            border-radius: 16px;
            padding: 14px 14px 6px;
            margin-bottom: 16px;
            border: 2px solid rgba(0,0,0,0.08);
        }}
        .team-title {{
            font-weight: 700;
            font-size: 16px;
            margin: 0 0 8px;
        }}
    footer {{
      text-align: center;
      color: var(--muted);
      font-size: 12px;
      padding: 0 0 24px;
    }}
  </style>
</head>
<body>
  <header>
                {logo_html}
        <div class=\"headline\">Campeonato de Espana por equipos 2026</div>
    <div class=\"subtitle\">Etapa 1 · 2026</div>
    <h1>Clasificacion de Equipos</h1>
  </header>
  <section class=\"grid\">
    {''.join(cards)}
  </section>
    <section class=\"table-section\">
        <div class=\"table-card\">
            <h2>Clasificacion por etapas</h2>
            <table>
                <thead>
                    <tr>
                        <th class=\"team-cell\">Equipo</th>
                        {''.join(f'<th>{name}</th>' for name in stage_names)}
                        <th>Total</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(
                            '<tr>'
                            f"<td class='team-cell'>{row['Equipo']}</td>" +
                            ''.join(f"<td>{value}</td>" for value in row['Etapas']) +
                            f"<td>{row['Total']}</td>"
                            '</tr>'
                            for row in classification_rows
                    )}
                </tbody>
            </table>
        </div>
        <div class=\"table-card\">
            <h2>Detalle por equipos</h2>
            {''.join(
                    f"<div class='team-block' style='background:{team_colors[idx % len(team_colors)]}'>"
                    f"<div class='team-title'>{group['Equipo']}</div>"
                    "<table><thead><tr>"
                    f"<th class='team-cell'>Jugador</th>" +
                    ''.join(f"<th>{name}</th>" for name in stage_names) +
                    "</tr></thead><tbody>" +
                    ''.join(
                            "<tr>" +
                            f"<td class='team-cell'>{row['Jugador']}</td>" +
                            ''.join(
                                    f"<td class='{'scored' if scored else ''}'>{score}</td>"
                                    for score, scored in zip(row['Etapas'], row['Scored'])
                            ) +
                            "</tr>"
                            for row in group['Rows']
                    ) +
                    "</tbody></table></div>"
                    for idx, group in enumerate(player_groups)
            )}
        </div>
    </section>
  <footer>Footgolf · Clasificacion por equipos</footer>
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


def build_pdf(results, output_path, classification_rows, player_groups, stage_names, logo_path):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Image,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ModuleNotFoundError:
        print("reportlab no esta instalado; se omite la generacion del PDF.")
        return

    styles = getSampleStyleSheet()
    body_style = styles["BodyText"]
    body_style.fontSize = 9

    max_team_len = max((len(str(item["Equipo"] or "")) for item in results), default=0)
    page_width = A4[0] - 36 * mm
    if max_team_len > 32:
        team_w = 90 * mm
        total_w = 30 * mm
        aportes_w = page_width - (12 * mm + team_w + total_w)
    else:
        team_w = 70 * mm
        total_w = 30 * mm
        aportes_w = page_width - (12 * mm + team_w + total_w)

    col_widths = [12 * mm, team_w, total_w, aportes_w]

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    flow = []
    if logo_path and logo_path.exists():
        flow.append(Image(str(logo_path), width=60 * mm, height=28 * mm))
        flow.append(Spacer(1, 4 * mm))
    flow.append(Paragraph("Campeonato de Espana por equipos 2026", styles["Title"]))
    flow.append(Paragraph("Clasificacion de Equipos - Etapa 1 2026", styles["Heading2"]))
    flow.append(Spacer(1, 6 * mm))

    data = [["#", "Equipo", "Total", "Aportes (golpes)"]]
    for idx, item in enumerate(results, start=1):
        total = f"{item['TotalGolpes']} golpes"
        aportes = str(item["GolpesPuntuaron"] or "")
        team = Paragraph(str(item["Equipo"]), body_style)
        data.append([str(idx), team, total, aportes])

    table = Table(data, colWidths=col_widths)
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1c2329")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c9d2d9")),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ]
    )

    style.add("BACKGROUND", (2, 1), (2, -1), colors.HexColor("#f4f0e6"))
    style.add("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold")
    style.add("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor("#1c2329"))
    style.add("ALIGN", (2, 1), (2, -1), "CENTER")

    style.add("BOX", (0, 1), (-1, 1), 1.5, colors.HexColor("#c8a23d"))
    style.add("BOX", (0, 2), (-1, 2), 1.5, colors.HexColor("#8a97a6"))
    style.add("BOX", (0, 3), (-1, 3), 1.5, colors.HexColor("#b26a4c"))
    style.add("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#fff6d5"))
    style.add("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#d9dee5"))
    style.add("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#f4e6db"))
    if len(data) > 4:
        style.add("BACKGROUND", (0, 4), (-1, -1), colors.white)

    style.add("TOPPADDING", (0, 1), (-1, -1), 6)
    style.add("BOTTOMPADDING", (0, 1), (-1, -1), 6)

    table.setStyle(style)
    flow.append(table)
    flow.append(Spacer(1, 8 * mm))

    team_name_style = ParagraphStyle(
        "team_name",
        parent=styles["BodyText"],
        fontSize=7,
        leading=8,
    )
    clas_header = ["Equipo"] + stage_names + ["Total"]
    clas_data = [clas_header]
    for row in classification_rows:
        team_cell = Paragraph(str(row["Equipo"]), team_name_style)
        clas_data.append([team_cell] + row["Etapas"] + [row["Total"]])

    clas_col_widths = [55 * mm] + [12 * mm] * len(stage_names) + [16 * mm]
    clas_table = Table(clas_data, colWidths=clas_col_widths)
    clas_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1c2329")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c9d2d9")),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
        ]
    )
    clas_table.setStyle(clas_style)
    flow.append(Paragraph("Clasificacion por etapas", styles["Heading3"]))
    flow.append(clas_table)
    flow.append(Spacer(1, 6 * mm))

    team_colors = [
        colors.HexColor("#f8e1b8"),
        colors.HexColor("#d6ecf4"),
        colors.HexColor("#f4d6e0"),
        colors.HexColor("#e1f1d2"),
        colors.HexColor("#f1e1c4"),
        colors.HexColor("#dfe2f6"),
        colors.HexColor("#f6e0d2"),
        colors.HexColor("#d2f0e9"),
    ]

    flow.append(Paragraph("Detalle por equipos", styles["Heading3"]))
    for idx, group in enumerate(player_groups):
        team_title = Table([[group["Equipo"]]], colWidths=[page_width])
        team_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), team_colors[idx % len(team_colors)]),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
        team_title.setStyle(team_style)
        flow.append(team_title)

        detail_header = ["Jugador"] + stage_names
        detail_data = [detail_header]
        scored_cells = []
        for row_idx, row in enumerate(group["Rows"], start=1):
            detail_data.append([row["Jugador"]] + row["Etapas"])
            for col_idx, scored in enumerate(row["Scored"], start=1):
                if scored:
                    scored_cells.append((col_idx, row_idx))

        detail_col_widths = [45 * mm] + [12 * mm] * len(stage_names)
        detail_table = Table(detail_data, colWidths=detail_col_widths)
        detail_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1c2329")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c9d2d9")),
                ("FONTSIZE", (0, 1), (-1, -1), 7),
            ]
        )
        for col_idx, row_idx in scored_cells:
            detail_style.add(
                "BACKGROUND",
                (col_idx, row_idx),
                (col_idx, row_idx),
                colors.HexColor("#c6efce"),
            )
        detail_table.setStyle(detail_style)
        flow.append(detail_table)
        flow.append(Spacer(1, 4 * mm))

    doc.build(flow)


def main():
    parser = argparse.ArgumentParser(
        description="Genera clasificacion por equipos y exporta HTML/PDF."
    )
    parser.add_argument(
        "--input-xlsx",
        default=str(BASE_DIR / "imports" / "etapa1_2026_equipos.xlsx"),
        help="Ruta al Excel con Equipos y Clasificacion etapa 1 2026.",
    )
    parser.add_argument(
        "--sheet-teams", default="Equipos", help="Nombre de la hoja de equipos."
    )
    parser.add_argument(
        "--sheet-scores",
        default="Clasificacion etapa 1 2026",
        help="Nombre de la hoja de clasificacion individual.",
    )
    parser.add_argument(
        "--output-html",
        default=str(BASE_DIR / "exports" / "clasificacion_equipos_etapa1_2026.html"),
        help="Ruta de salida HTML.",
    )
    parser.add_argument(
        "--output-pdf",
        default=str(BASE_DIR / "exports" / "clasificacion_equipos_etapa1_2026.pdf"),
        help="Ruta de salida PDF.",
    )
    parser.add_argument(
        "--update-xlsx",
        action="store_true",
        help="Actualiza la hoja 'Clasificacion equipos' en el Excel de entrada.",
    )
    args = parser.parse_args()

    input_path = Path(args.input_xlsx)
    output_html = Path(args.output_html)
    output_pdf = Path(args.output_pdf)
    output_html.parent.mkdir(parents=True, exist_ok=True)
    output_pdf.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(input_path, data_only=True)
    ws_teams = wb[args.sheet_teams]
    ws_scores = wb[args.sheet_scores]

    teams = read_teams(ws_teams)
    scores = read_scores(ws_scores)
    results = compute_results(teams, scores)
    stage_names = [f"Etapa {i}" for i in range(1, 9)]
    classification_rows = compute_team_points(results, len(stage_names))
    player_groups = build_player_groups(teams, scores, len(stage_names))

    if args.update_xlsx:
        write_clasificacion_sheet(wb, results)
        wb.save(input_path)

    build_html(results, output_html, classification_rows, player_groups, stage_names, LOGO_PATH)
    build_pdf(results, output_pdf, classification_rows, player_groups, stage_names, LOGO_PATH)

    print(str(output_html))
    print(str(output_pdf))


if __name__ == "__main__":
    main()
